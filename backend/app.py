from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
import time
from datetime import datetime
import os
from dotenv import load_dotenv
import hashlib
import base64
import stat
from pathlib import Path
import sys
import webbrowser
from urllib.parse import quote
import pyautogui

# Patch pywhatkit's internet check before importing it
import types
from importlib import import_module

# Import the core module where check_connection is defined
pywhatkit_core = import_module('pywhatkit.core.core')

# Create a patched version of check_connection that doesn't require internet
def patched_check_connection():
    """Patched version that bypasses internet connectivity check"""
    pass

# Replace the original function with our patched version
pywhatkit_core.check_connection = patched_check_connection

# Now we can safely import pywhatkit
import pywhatkit

# Add these constants at the top of the file after imports
ASSETS_DIR = os.path.join(os.path.dirname(__file__), 'assets')
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), 'uploads')
ATTACHMENT_BUTTON = os.path.join(ASSETS_DIR, 'attachment_button.png')
IMAGE_BUTTON = os.path.join(ASSETS_DIR, 'image_button.png')

app = Flask(__name__)
CORS(app)

# Maximum file sizes (in bytes)
MAX_EXCEL_SIZE = 5 * 1024 * 1024  # 5MB for Excel files
MAX_IMAGE_SIZE = 2 * 1024 * 1024  # 2MB for images

# Configure maximum content length for the entire request
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10MB total request size

# Keep track of sent message hashes to avoid duplicates
sent_messages = set()

def set_full_permissions(path):
    """Set full read/write permissions for everyone on the given path"""
    try:
        if sys.platform == 'win32':
            import win32security
            import ntsecuritycon as con
            
            sd = win32security.GetFileSecurity(path, win32security.DACL_SECURITY_INFORMATION)
            dacl = win32security.ACL()
            everyone = win32security.ConvertStringSidToSid("S-1-1-0")
            dacl.AddAccessAllowedAce(win32security.ACL_REVISION, con.FILE_ALL_ACCESS, everyone)
            sd.SetSecurityDescriptorDacl(1, dacl, 0)
            win32security.SetFileSecurity(path, win32security.DACL_SECURITY_INFORMATION, sd)
        else:
            os.chmod(path, stat.S_IRWXU | stat.S_IRWXG | stat.S_IRWXO)
        return True
    except Exception as e:
        print(f"Error setting permissions for {path}: {str(e)}")
        return False

def ensure_directory_with_permissions(directory):
    """Create directory if it doesn't exist and set proper permissions"""
    try:
        Path(directory).mkdir(parents=True, exist_ok=True)
        set_full_permissions(directory)
        test_file = os.path.join(directory, 'test_permissions.txt')
        try:
            with open(test_file, 'w') as f:
                f.write('test')
            os.remove(test_file)
            return True
        except Exception as e:
            print(f"Permission test failed for {directory}: {str(e)}")
            return False
    except Exception as e:
        print(f"Error creating directory {directory}: {str(e)}")
        return False

# Setup directories and permissions
print("Setting up directories and permissions...")
required_dirs = [
    os.path.join(os.path.dirname(__file__), 'uploads'),
    os.path.join(os.path.dirname(__file__), 'assets')
]

for directory in required_dirs:
    if not ensure_directory_with_permissions(directory):
        print(f"WARNING: Could not set up {directory} properly")

# Load environment variables
load_dotenv()

def send_message_automated(phone, message, image_path=None):
    try:
        # Format the phone number
        phone = str(phone).strip()
        # Remove any non-digit characters except '+'
        phone = ''.join(char for char in phone if char.isdigit() or char == '+')
        
        # If number doesn't start with '+', assume it's a local number
        if not phone.startswith('+'):
            phone = '+' + phone

        if image_path:
            # Using optimized settings for faster image sending with single tab
            pywhatkit.sendwhats_image(
                receiver=phone,
                img_path=image_path,
                caption=message,
                wait_time=15,  # Increased wait time to ensure proper loading
                tab_close=False,  # Keep tab open for next message
                close_time=3  # Slightly longer close time to ensure message sends
            )
            # Ensure positive sleep time
            time.sleep(3)
        else:
            # For text-only messages
            pywhatkit.sendwhatmsg_instantly(
                phone_no=phone,
                message=message,
                wait_time=15,  # Increased wait time for stability
                tab_close=False,  # Keep tab open for next message
                close_time=3
            )
            # Ensure positive sleep time
            time.sleep(3)

        return True, None

    except Exception as e:
        if "sleep length must be non-negative" in str(e):
            # Add extra delay and retry once
            time.sleep(5)
            try:
                if image_path:
                    pywhatkit.sendwhats_image(
                        receiver=phone,
                        img_path=image_path,
                        caption=message,
                        wait_time=15,
                        tab_close=False,
                        close_time=3
                    )
                else:
                    pywhatkit.sendwhatmsg_instantly(
                        phone_no=phone,
                        message=message,
                        wait_time=15,
                        tab_close=False,
                        close_time=3
                    )
                return True, None
            except Exception as retry_error:
                return False, str(retry_error)
        return False, str(e)

@app.route('/send-whatsapp', methods=['POST'])
def send_whatsapp():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        message = request.form.get('message')
        image_data = request.form.get('image')
        
        # Check Excel file size
        if file.content_length and file.content_length > MAX_EXCEL_SIZE:
            return jsonify({'error': f'Excel file size exceeds maximum limit of {MAX_EXCEL_SIZE/1024/1024}MB'}), 413

        # Check image size if present
        if image_data and ',' in image_data:
            image_size = len(image_data.split(',')[1]) * 3/4  # Base64 size to actual size
            if image_size > MAX_IMAGE_SIZE:
                return jsonify({'error': f'Image size exceeds maximum limit of {MAX_IMAGE_SIZE/1024/1024}MB'}), 413

        if not message:
            return jsonify({'error': 'No message provided'}), 400

        # Handle image if provided
        image_path = None
        if image_data and ',' in image_data:
            try:
                # Extract the base64 data
                image_base64 = image_data.split(',')[1]
                # Generate a unique filename using uuid
                import uuid
                image_filename = f"whatsapp_image_{uuid.uuid4().hex}.png"
                image_path = os.path.join(UPLOAD_DIR, image_filename)
                
                # Ensure uploads directory exists and has proper permissions
                if not ensure_directory_with_permissions(UPLOAD_DIR):
                    return jsonify({'error': 'Failed to set up uploads directory'}), 500
                
                # Save the image with proper error handling
                try:
                    with open(image_path, "wb") as f:
                        f.write(base64.b64decode(image_base64))
                    
                    # Set permissions on the new file
                    if not set_full_permissions(image_path):
                        return jsonify({'error': 'Failed to set proper permissions on uploaded file'}), 500
                        
                except Exception as e:
                    return jsonify({'error': f'Failed to save image: {str(e)}'}), 500

            except Exception as e:
                return jsonify({'error': f'Error processing image: {str(e)}'}), 500

        # Create a hash of the message to check for duplicates
        message_hash = hashlib.md5(message.encode()).hexdigest()
        
        # Check if this exact message has been sent before
        if message_hash in sent_messages:
            return jsonify({
                'success': False,
                'error': 'This exact message has already been sent. Please modify the message to send again.'
            }), 400

        # Read Excel file using openpyxl
        workbook = load_workbook(filename=file)
        worksheet = workbook.active
        
        # Find the phone column
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        try:
            phone_col_idx = header_row.index('phone')
        except ValueError:
            return jsonify({'error': 'Excel file must contain a "phone" column'}), 400

        results = []
        success_count = 0
        # Track processed phone numbers in this batch
        processed_phones = set()

        try:
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                phone = str(row[phone_col_idx]).strip()
                
                # Skip if this phone number has already been processed in this batch
                if phone in processed_phones:
                    results.append({
                        'phone': phone,
                        'status': 'skipped',
                        'error': 'Phone number already processed in this batch'
                    })
                    continue
                
                try:
                    success, error = send_message_automated(phone, message, image_path)
                    
                    if success:
                        success_count += 1
                        results.append({
                            'phone': phone,
                            'status': 'success',
                            'message_sid': 'N/A',
                            'delivery_status': 'sent',
                            'error': None
                        })
                        print(f"Message sent to {phone}")
                        # Add to processed phones after successful send
                        processed_phones.add(phone)
                    else:
                        results.append({
                            'phone': phone,
                            'status': 'failed',
                            'error': error
                        })
                        print(f"Error sending to {phone}: {error}")
                    
                    # Reduced delay between messages from 2 seconds to 1 second
                    time.sleep(1)
                    
                except Exception as e:
                    print(f"Error sending to {phone}: {str(e)}")
                    results.append({
                        'phone': phone,
                        'status': 'failed',
                        'error': str(e)
                    })

        finally:
            # Clean up the image file if it exists
            if image_path and os.path.exists(image_path):
                try:
                    os.remove(image_path)
                except Exception as e:
                    print(f"Error removing temporary image file: {str(e)}")

        # Only store the message hash if at least one message was sent successfully
        if success_count > 0:
            sent_messages.add(message_hash)

        return jsonify({
            'success': True,
            'results': results,
            'note': 'Please ensure you are logged into WhatsApp Web before sending messages'
        })

    except Exception as e:
        print(f"General error: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

if __name__ == '__main__':
    app.run(debug=True, port=5000)